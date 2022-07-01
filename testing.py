import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import time


sheetMade = False
#spreadsheet setup
if os.path.isfile('./zoneTesting.xlsx'):
	wb = load_workbook('zoneTesting.xlsx')
	ws = wb.active

else:
	wb = Workbook()
	ws = [wb.active]
	ws[0].title = 'Zone 1'
	for i in range(25):
		ws[i].cell(row=1, column=1, value="Trial Number")
		ws[i].cell(row=33, column=1, value="Num Correct")
		ws[i].cell(row=34, column=1, value="Percentage")
		if i < 24:
			ws.append(wb.create_sheet("Zone " + str(i + 2)))
if 'Zone 25' in wb.sheetnames:
	nothing = 0
else:
	wb.create_sheet("Zone 25")
	wa = wb.get_sheet_by_name('Zone 25')
	wa.cell(row=1, column=1, value="Trial Number")
	wa.cell(row=33, column=1, value="Num Correct")
	wa.cell(row=34, column=1, value="Percentage")




#Check if all avgs are added
check = True
systemAvgSum = 0.0
for i in range(25):
	ws1 = wb.get_sheet_by_name('Zone ' + str(i+1))
	if str(ws1['B34'].value) != "None":
		print("Zone " + str(i + 1) + " " +str(ws1['B34'].value))
		systemAvgSum = systemAvgSum + float(ws1['B34'].value)
		x = float(ws1['B34'].value)
	else:
		check = False

if check:
	avg =  float(systemAvgSum)/25.0
	print("all data was entered and the the system average is: " + str(avg))

#zone input
zoneIn = input("Enter Zone Number:")
print("Zone selected : " + zoneIn)
zoneNum = int(zoneIn)

#api setup
headers = {'Authorization': 'Token token="NaUtIlUs"',}
url = 'http://tss-14127/api/tents/'


#time check
now = datetime.now()
current_time = now.strftime("%H:%M")
currentHour = int(current_time[0:2])
currentMinute = int(current_time[3:5])
print("current time is: " + str(current_time))
print("test")

#Active tent count
j = 0
numCorrect = 0;


#ws[0].cell(row=1, column=zoneNum + 1, value= zoneNum)
ws1 = wb.get_sheet_by_name('Zone ' + str(zoneNum))

tent_aval = [1,2,4,5,7,8,9,10,11,12,13,14,15,16,17,19,20,21,22,23,25,26,27,28,29,30,31,32,33,52]

#gets Zone
for k in range(24):
	ws1.cell(row = 1,column= k + 2, value = ("Trial " + str(k+1)) )
	now = datetime.now()
	current_time = now.strftime("%H:%M:%S")
	currentHour = int(current_time[0:2])
	currentMinute = int(current_time[3:5])
	currentSecond = int(current_time[6:8])
	for i in range(30):
		#api call
		fullUrl = url + str(tent_aval[i])
		response = requests.get(fullUrl, headers=headers)
		tentInfo = response.json()

		#parse
		print("tent number: " + str(tent_aval[i]))
		tent = tentInfo['tent']
		tentNum = tent['number']
		scanTime = tent['scanned_at']
		scanTime = scanTime[11:20]
		hour = int(scanTime[0:2]) -4
		minute = int(scanTime[3:5])
		second = int(scanTime[6:8])
		zone = int(tent['location_number'])
		#active?

		
		if currentSecond  <= 7 and second > 52:  #check minute rollover in this case system may think a tent is off
			currentMinuteR = currentMinute - 1
		elif second  <= 5 and currentSecond > 52:
			currentMinuteR = currentMinute + 1
		else:
			currentMinuteR = currentMinute

		if currentHour == hour and currentMinuteR == minute:
			j = j + 1
		else:
			print("FAILURE! A tent has gone to sleep and no data will be recorded. Please restart the program")
			print("Current time : " + current_time + " " + str(currentHour)+str(currentMinute))
			print("tent time: " + str(hour) + ":" +  str(minute))
			print("tent number: " + str(tentNum) )
			quit()

		if zoneNum == zone:
			numCorrect = numCorrect + 1
		print("Tent Number " + str(tentNum) + " Zone Number " + str(zone) + " Scan Time " +str(scanTime))

		ws1.cell(row=i+2, column= k + 2, value=zone)
		
	time.sleep(5)


		
percent = float(numCorrect)/720.0
print("Number correct: " + str(numCorrect))
ws1.cell(row = 34, column = 2, value = percent)
ws1.cell(row=33, column= 2, value=numCorrect)
wb.save('zoneTesting.xlsx')
print("sucess")














